import os
from typing import List

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from RPA.Excel.Files import Files

from util.error_manager import ErrorManager


class ExcelGenerator:
    """
    Class responsible for generating excel
    files.
    """

    def __init__(self) -> None:
        self.file = Files()

    def add_data(self, data: List) -> None:
        """
        Method responsible for adding
        data inside the excel file.
        """
        try:
            self.file.append_rows_to_worksheet(data)
        except Exception as error:
            message = f"Error adding data: {error}"
            error_code = 11

            raise ErrorManager(message, error_code)

    def adjust_width_to_show_data(self, excel_path: str) -> None:
        """
        Method responsible for defining the width of the columns.
        """
        try:
            work_book = load_workbook(excel_path)
            work_sheet = work_book.active

            for column in work_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                adjusted_width = max_length + 2
                work_sheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

            work_book.save(excel_path)
        except Exception as error:
            message = f"Error adjusting width: {error}"
            error_code = 12

            raise ErrorManager(message, error_code)

    def add_image_to_cell(
        self, image_path: str, cell: str, excel_path: str
    ) -> bool:
        """
        Method responsible for adding image to a
        cell in a excel file.
        """
        try:
            if os.path.isfile(image_path) and os.path.isfile(excel_path):
                work_book = load_workbook(excel_path)
                work_sheet = work_book.active
                image = Image(image_path)

                work_sheet.add_image(image, cell)
                work_book.save(excel_path)

                return True

            return False
        except Exception as error:
            message = f"Error adding image to cell: {error}"
            error_code = 13

            raise ErrorManager(message, error_code)

    def create_workbook(self) -> None:
        """
        Method responsible for creating
        a workbook.
        """
        try:
            self.file.create_workbook()
        except Exception as error:
            message = f"Error creating workbook: {error}"
            error_code = 14

            raise ErrorManager(message, error_code)

    def save_workbook(self, path: str) -> None:
        """
        Method responsible for saving
        the workbook.
        """
        try:
            self.file.save_workbook(path)
        except Exception as error:
            message = f"Error creating workbook: {error}"
            error_code = 15

            raise ErrorManager(message, error_code)

    def set_row_height(self, excel_path: str, height: float) -> None:
        """
        Method responsible for defining the height of the lines.
        """
        try:
            work_book = load_workbook(excel_path)
            work_sheet = work_book.active

            for row in work_sheet.iter_rows(min_row=2):
                work_sheet.row_dimensions[row[0].row].height = height

            work_book.save(excel_path)
        except Exception as error:
            message = f"Error setting row height: {error}"
            error_code = 16

            raise ErrorManager(message, error_code)
